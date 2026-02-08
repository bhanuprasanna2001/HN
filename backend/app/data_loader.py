"""Load and parse the SupportMind Excel workbook into structured dicts."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# Sheet names we care about
SHEET_NAMES = [
    "Conversations",
    "Tickets",
    "Questions",
    "Scripts_Master",
    "Placeholder_Dictionary",
    "Knowledge_Articles",
    "Existing_Knowledge_Articles",
    "KB_Lineage",
    "Learning_Events",
    "QA_Evaluation_Prompt",
]


def _parse_sheet(ws: Any) -> list[dict[str, Any]]:
    """Convert a worksheet into a list of dicts keyed by header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record = {}
        has_data = False
        for key, val in zip(headers, row):
            if val is not None:
                has_data = True
            record[key] = _serialize(val)
        if has_data:
            records.append(record)
    return records


def _serialize(val: Any) -> Any:
    """Convert Excel cell values to JSON-safe types."""
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return val


def load_workbook_data(path: str | Path) -> dict[str, list[dict[str, Any]]]:
    """Load the entire workbook and return {sheet_name: [records]}."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Data file not found: {path}")

    logger.info("Loading workbook from %s", path)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    data: dict[str, list[dict[str, Any]]] = {}
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            data[name] = _parse_sheet(wb[name])
            logger.info("  %s: %d records", name, len(data[name]))
        else:
            logger.warning("  Sheet '%s' not found in workbook", name)
            data[name] = []

    # Extract QA rubric as a single string from the first cell
    if "QA_Evaluation_Prompt" in wb.sheetnames:
        qa_ws = wb["QA_Evaluation_Prompt"]
        first_row = next(qa_ws.iter_rows(max_row=1, values_only=True), None)
        data["_qa_rubric"] = str(first_row[0]) if first_row and first_row[0] else ""
    else:
        data["_qa_rubric"] = ""

    wb.close()
    return data


def build_ticket_lookup(data: dict) -> dict[str, dict]:
    """Build a {Ticket_Number: record} lookup from Tickets sheet."""
    return {r["Ticket_Number"]: r for r in data.get("Tickets", []) if r.get("Ticket_Number")}


def build_conversation_lookup(data: dict) -> dict[str, dict]:
    """Build a {Ticket_Number: conversation} lookup from Conversations sheet."""
    return {r["Ticket_Number"]: r for r in data.get("Conversations", []) if r.get("Ticket_Number")}


def build_script_lookup(data: dict) -> dict[str, dict]:
    """Build a {Script_ID: record} lookup from Scripts_Master sheet."""
    return {r["Script_ID"]: r for r in data.get("Scripts_Master", []) if r.get("Script_ID")}


def build_kb_lookup(data: dict) -> dict[str, dict]:
    """Build a {KB_Article_ID: record} lookup from Knowledge_Articles sheet."""
    return {r["KB_Article_ID"]: r for r in data.get("Knowledge_Articles", []) if r.get("KB_Article_ID")}
